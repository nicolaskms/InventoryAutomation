# backend/src/server.py
from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from io import BytesIO
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import re
from typing import List, Tuple, Optional, Any
import logging

# IMPORTA suas funções já existentes do módulo compare.py
from compare import carregar_planilha, comparar, _extrai_local

# Importa a função gerar_em_branco caso exista em blank.py (opcional)
try:
    from blank import gerar_em_branco  # type: ignore
except Exception:
    gerar_em_branco = None  # pode não existir; /blank ficará disponível só se presente

logger = logging.getLogger("uvicorn.error")

app = FastAPI(title="InventoryAutomation API")

# cria DATA_DIR para arquivos temporários (compatível com outras versões)
DATA_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data"))
os.makedirs(DATA_DIR, exist_ok=True)

# Libera o front (Vite)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

_rx_loc = re.compile(r"^([A-Za-z]+)?(\d+)?([A-Za-z]+)?$")


def _sort_key_gaveta(s: str) -> Tuple[str, int, str]:
    s = str(s).strip()
    m = _rx_loc.match(s) or None
    if not m:
        return (s.lower(), 0, "")
    g1, g2, g3 = m.groups()
    letra = (g1 or "").lower()
    numero = int(g2) if g2 and g2.isdigit() else 0
    sufx = (g3 or "").lower()
    return (letra, numero, sufx)


def _auto_fit_and_center(xlsx_bytes: BytesIO, sheet_name: Optional[str] = None) -> BytesIO:
    xlsx_bytes.seek(0)
    wb = load_workbook(xlsx_bytes)
    ws = wb[sheet_name] if (sheet_name and sheet_name in wb.sheetnames) else wb.active

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = center

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[letter].width = min(max(12, int(max_len * 1.2)), 60)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def _df_to_xlsx_bytes(df: pd.DataFrame, sheet_name: str = "Relatorio") -> BytesIO:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    buf.seek(0)
    return buf


async def _first_uploadfile_from_request(request: Request) -> Optional[UploadFile]:
    """
    Utility: pega o primeiro UploadFile presente no multipart/form-data
    independente do nome do campo (flexibilidade para o frontend).
    Usa uma checagem flexível (não depende estritamente do tipo).
    """
    form = await request.form()
    for k, v in form.items():
        # v pode ser UploadFile (Starlette) ou outro objeto com .filename
        if hasattr(v, "filename") and getattr(v, "filename"):
            logger.debug("Found upload field '%s' filename='%s'", k, getattr(v, "filename"))
            return v  # tipo é UploadFile/Starlette UploadFile
    return None


@app.post("/compare")
async def compare_endpoint(
    planilha_oficial: UploadFile = File(...),
    planilha_divergente: UploadFile = File(...),
):
    try:
        oficial_bytes = BytesIO(await planilha_oficial.read())
        divergente_bytes = BytesIO(await planilha_divergente.read())

        df_oficial = carregar_planilha(oficial_bytes)
        df_div = carregar_planilha(divergente_bytes)

        df_out: pd.DataFrame = comparar(df_oficial, df_div)

        buf = _df_to_xlsx_bytes(df_out, sheet_name="Relatorio")
        buf = _auto_fit_and_center(buf, sheet_name="Relatorio")

        filename = "relatorio_auditoria_comparacao.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except Exception as e:
        logger.exception("Error in /compare")
        raise HTTPException(status_code=400, detail=f"Erro ao processar: {e}")


@app.post("/blind-template")
async def blind_template(
    request: Request,
    planilha_oficial: UploadFile = File(None),
):
    """
    Recebe uma planilha WMS (aceita campo 'planilha_oficial' ou qualquer arquivo multipart)
    e devolve um XLSX 'às cegas' com apenas a coluna 'gaveta' preenchida.
    """
    try:
        # Primeiro tenta o UploadFile explicitamente nomeado
        planilha = planilha_oficial if (planilha_oficial and getattr(planilha_oficial, "filename", None)) else None

        # Se não veio via parâmetro, tenta pegar o primeiro arquivo do form (fallback)
        if not planilha:
            planilha = await _first_uploadfile_from_request(request)

        if not planilha or not getattr(planilha, "filename", None):
            logger.warning("No file found in request for /blind-template. Form keys: %s", list((await request.form()).keys()))
            raise HTTPException(status_code=400, detail="Arquivo inválido")

        logger.info("Received file for blind-template: %s", getattr(planilha, "filename"))

        oficial_bytes = BytesIO(await planilha.read())
        df_wms = carregar_planilha(oficial_bytes)

        if "gaveta" not in df_wms.columns:
            df_wms["gaveta"] = ""

        gavetas: List[str] = (
            df_wms["gaveta"]
            .astype(str)
            .map(lambda x: _extrai_local(x).strip())
            .replace("", pd.NA)
            .dropna()
            .unique()
            .tolist()
        )

        gavetas_ordenadas = sorted(gavetas, key=_sort_key_gaveta)

        cols = ["gaveta", "cod", "produto", "lote", "quantidade", "observacao"]
        df_out = pd.DataFrame({"gaveta": gavetas_ordenadas})
        for c in cols[1:]:
            df_out[c] = ""

        buf = _df_to_xlsx_bytes(df_out, sheet_name="Relatorio_As_Cegas")
        buf = _auto_fit_and_center(buf, sheet_name="Relatorio_As_Cegas")

        filename = "relatorio_as_cegas.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except HTTPException:
        # repropaga HTTPException sem alteração
        raise
    except Exception as e:
        logger.exception("Error in /blind-template")
        raise HTTPException(status_code=400, detail=f"Erro ao gerar relatório às cegas: {e}")


@app.post("/blank")
async def blank_endpoint(request: Request):
    if gerar_em_branco is None:
        raise HTTPException(status_code=404, detail="Endpoint /blank não disponível (blank.gerar_em_branco ausente)")

    try:
        wms = await _first_uploadfile_from_request(request)
        if not wms or not getattr(wms, "filename", None):
            raise HTTPException(status_code=400, detail="Arquivo inválido")

        wms_path = os.path.join(DATA_DIR, "wms_upload.xlsx")
        with open(wms_path, "wb") as f:
            f.write(await wms.read())

        df_blank = gerar_em_branco(wms_path)

        buf = _df_to_xlsx_bytes(df_blank, sheet_name="relatorio")
        buf = _auto_fit_and_center(buf, sheet_name="relatorio")

        filename = "relatorio_em_branco.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    except Exception as e:
        logger.exception("Error in /blank")
        raise HTTPException(status_code=400, detail=f"Erro ao processar planilha: {e}")